"""
Сервис для экспорта данных в различных форматах
"""
import csv
import json
import asyncio
from io import StringIO, BytesIO
from datetime import datetime
from typing import Dict, List, Any, Union
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from database.repositories.user_repository import UserRepository
from database.repositories.team_repository import TeamRepository
from database.repositories.tournament_repository import TournamentRepository


class ExportService:
    """Сервис для экспорта данных"""
    
    def __init__(self):
        # Репозитории используют статические методы, экземпляры не нужны
        pass
    
    async def export_users_csv(self) -> StringIO:
        """Экспорт пользователей в CSV"""
        output = StringIO()
        writer = csv.writer(output, delimiter=';')
        
        # Заголовки
        headers = [
            'ID', 'Telegram ID', 'Имя пользователя', 'Полное имя', 
            'Роль', 'Регион', 'Язык', 'Заблокирован',
            'Дата регистрации', 'Последнее обновление'
        ]
        writer.writerow(headers)
        
        # Получаем всех пользователей
        users = await UserRepository.get_all_users(limit=None)
        
        for user in users:
            row = [
                user.id,
                user.telegram_id,
                user.username or '',
                user.full_name or '',
                user.role or 'user',
                user.region or '',
                user.language or '',
                'Да' if user.is_blocked else 'Нет',
                user.created_at.strftime('%Y-%m-%d %H:%M:%S') if user.created_at else '',
                user.updated_at.strftime('%Y-%m-%d %H:%M:%S') if user.updated_at else ''
            ]
            writer.writerow(row)
        
        output.seek(0)
        return output
    
    async def export_teams_csv(self) -> StringIO:
        """Экспорт команд в CSV"""
        output = StringIO()
        writer = csv.writer(output, delimiter=';')
        
        # Заголовки
        headers = [
            'ID', 'Название', 'Описание', 'Капитан ID', 'Капитан',
            'Статус', 'Максимум участников', 'Текущих участников',
            'Дата создания', 'Последнее обновление'
        ]
        writer.writerow(headers)
        
        # Получаем все команды
        teams = await TeamRepository.get_all_teams()
        
        for team in teams:
            # Подсчитываем участников
            members_count = len(team.players) if hasattr(team, 'players') and team.players else 0
            
            row = [
                team.id,
                team.name,
                team.description or '',
                team.captain_id or '',
                team.captain.full_name if team.captain else '',
                team.status or 'pending',
                team.max_members or '',
                members_count,
                team.created_at.strftime('%Y-%m-%d %H:%M:%S') if team.created_at else '',
                team.updated_at.strftime('%Y-%m-%d %H:%M:%S') if team.updated_at else ''
            ]
            writer.writerow(row)
        
        output.seek(0)
        return output
    
    async def export_tournaments_csv(self) -> StringIO:
        """Экспорт турниров в CSV"""
        output = StringIO()
        writer = csv.writer(output, delimiter=';')
        
        # Заголовки
        headers = [
            'ID', 'Название', 'Описание', 'Статус', 'Максимум команд',
            'Зарегистрированных команд', 'Дата начала', 'Дата окончания',
            'Дата создания', 'Последнее обновление'
        ]
        writer.writerow(headers)
        
        # Получаем все турниры
        tournaments = await TournamentRepository.get_all_tournaments(limit=None)
        
        for tournament in tournaments:
            # Подсчитываем зарегистрированные команды
            registered_count = len([r for r in tournament.registrations if r.status == 'approved'])
            
            row = [
                tournament.id,
                tournament.name,
                tournament.description or '',
                tournament.status or 'draft',
                tournament.max_teams or '',
                registered_count,
                tournament.start_date.strftime('%Y-%m-%d %H:%M:%S') if tournament.start_date else '',
                tournament.end_date.strftime('%Y-%m-%d %H:%M:%S') if tournament.end_date else '',
                tournament.created_at.strftime('%Y-%m-%d %H:%M:%S') if tournament.created_at else '',
                tournament.updated_at.strftime('%Y-%m-%d %H:%M:%S') if tournament.updated_at else ''
            ]
            writer.writerow(row)
        
        output.seek(0)
        return output
    
    async def export_users_json(self) -> str:
        """Экспорт пользователей в JSON"""
        users = await UserRepository.get_all_users(limit=None)
        
        users_data = []
        for user in users:
            user_data = {
                'id': user.id,
                'telegram_id': user.telegram_id,
                'username': user.username,
                'full_name': user.full_name,
                'role': user.role,
                'region': user.region,
                'language': user.language,
                'is_blocked': user.is_blocked,
                'created_at': user.created_at.isoformat() if user.created_at else None,
                'updated_at': user.updated_at.isoformat() if user.updated_at else None
            }
            users_data.append(user_data)
        
        return json.dumps(users_data, ensure_ascii=False, indent=2)
    
    async def export_teams_json(self) -> str:
        """Экспорт команд в JSON"""
        teams = await TeamRepository.get_all_teams()
        
        teams_data = []
        for team in teams:
            players = [
                {
                    'id': player.id,
                    'nickname': player.nickname,
                    'game_id': player.game_id,
                    'is_substitute': player.is_substitute,
                    'position': player.position
                }
                for player in (team.players if hasattr(team, 'players') and team.players else [])
            ]
            
            team_data = {
                'id': team.id,
                'name': team.name,
                'description': team.description,
                'status': team.status,
                'max_members': team.max_members,
                'captain': {
                    'id': team.captain.id,
                    'full_name': team.captain.full_name
                } if team.captain else None,
                'players': players,
                'created_at': team.created_at.isoformat() if team.created_at else None,
                'updated_at': team.updated_at.isoformat() if team.updated_at else None
            }
            teams_data.append(team_data)
        
        return json.dumps(teams_data, ensure_ascii=False, indent=2)
    
    async def export_tournaments_json(self) -> str:
        """Экспорт турниров в JSON"""
        tournaments = await TournamentRepository.get_all_tournaments(limit=None)
        
        tournaments_data = []
        for tournament in tournaments:
            registrations = [
                {
                    'team_id': reg.team_id,
                    'team_name': reg.team.name if reg.team else None,
                    'status': reg.status,
                    'registered_at': reg.created_at.isoformat() if reg.created_at else None
                }
                for reg in tournament.registrations
            ]
            
            tournament_data = {
                'id': tournament.id,
                'name': tournament.name,
                'description': tournament.description,
                'status': tournament.status,
                'max_teams': tournament.max_teams,
                'start_date': tournament.start_date.isoformat() if tournament.start_date else None,
                'end_date': tournament.end_date.isoformat() if tournament.end_date else None,
                'registrations': registrations,
                'created_at': tournament.created_at.isoformat() if tournament.created_at else None,
                'updated_at': tournament.updated_at.isoformat() if tournament.updated_at else None
            }
            tournaments_data.append(tournament_data)
        
        return json.dumps(tournaments_data, ensure_ascii=False, indent=2)
    
    async def export_excel(self) -> BytesIO:
        """Экспорт всех данных в Excel"""
        wb = Workbook()
        
        # Стили
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        
        # Удаляем стандартный лист
        wb.remove(wb.active)
        
        # Экспорт пользователей
        await self._add_users_sheet(wb, header_font, header_fill)
        
        # Экспорт команд
        await self._add_teams_sheet(wb, header_font, header_fill)
        
        # Экспорт турниров
        await self._add_tournaments_sheet(wb, header_font, header_fill)
        
        # Статистика
        await self._add_statistics_sheet(wb, header_font, header_fill)
        
        # Сохраняем в BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        return output
    
    async def _add_users_sheet(self, wb: Workbook, header_font: Font, header_fill: PatternFill):
        """Добавляет лист с пользователями"""
        ws = wb.create_sheet("Пользователи")
        
        # Заголовки
        headers = [
            'ID', 'Telegram ID', 'Имя пользователя', 'Полное имя', 
            'Роль', 'Регион', 'Язык', 'Заблокирован', 'Дата регистрации'
        ]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
        
        # Данные
        users = await UserRepository.get_all_users(limit=None)
        for row, user in enumerate(users, 2):
            ws.cell(row=row, column=1, value=user.id)
            ws.cell(row=row, column=2, value=user.telegram_id)
            ws.cell(row=row, column=3, value=user.username or '')
            ws.cell(row=row, column=4, value=user.full_name or '')
            ws.cell(row=row, column=5, value=user.role or 'user')
            ws.cell(row=row, column=6, value=user.region or '')
            ws.cell(row=row, column=7, value=user.language or '')
            ws.cell(row=row, column=8, value='Да' if user.is_blocked else 'Нет')
            ws.cell(row=row, column=9, value=user.created_at)
        
        # Автоширина столбцов
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].auto_size = True
    
    async def _add_teams_sheet(self, wb: Workbook, header_font: Font, header_fill: PatternFill):
        """Добавляет лист с командами"""
        ws = wb.create_sheet("Команды")
        
        # Заголовки
        headers = [
            'ID', 'Название', 'Описание', 'Капитан', 'Статус',
            'Максимум участников', 'Текущих участников', 'Дата создания'
        ]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
        
        # Данные
        teams = await TeamRepository.get_all_teams()
        for row, team in enumerate(teams, 2):
            members_count = len(team.players) if hasattr(team, 'players') and team.players else 0
            
            ws.cell(row=row, column=1, value=team.id)
            ws.cell(row=row, column=2, value=team.name)
            ws.cell(row=row, column=3, value=team.description or '')
            ws.cell(row=row, column=4, value=team.captain.full_name if team.captain else '')
            ws.cell(row=row, column=5, value=team.status or 'pending')
            ws.cell(row=row, column=6, value=team.max_members or '')
            ws.cell(row=row, column=7, value=members_count)
            ws.cell(row=row, column=8, value=team.created_at)
        
        # Автоширина столбцов
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].auto_size = True
    
    async def _add_tournaments_sheet(self, wb: Workbook, header_font: Font, header_fill: PatternFill):
        """Добавляет лист с турнирами"""
        ws = wb.create_sheet("Турниры")
        
        # Заголовки
        headers = [
            'ID', 'Название', 'Описание', 'Статус', 'Максимум команд',
            'Зарегистрированных команд', 'Дата начала', 'Дата окончания', 'Дата создания'
        ]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
        
        # Данные
        tournaments = await TournamentRepository.get_all_tournaments(limit=None)
        for row, tournament in enumerate(tournaments, 2):
            registered_count = len([r for r in tournament.registrations if r.status == 'approved'])
            
            ws.cell(row=row, column=1, value=tournament.id)
            ws.cell(row=row, column=2, value=tournament.name)
            ws.cell(row=row, column=3, value=tournament.description or '')
            ws.cell(row=row, column=4, value=tournament.status or 'draft')
            ws.cell(row=row, column=5, value=tournament.max_teams or '')
            ws.cell(row=row, column=6, value=registered_count)
            ws.cell(row=row, column=7, value=tournament.start_date)
            ws.cell(row=row, column=8, value=tournament.end_date)
            ws.cell(row=row, column=9, value=tournament.created_at)
        
        # Автоширина столбцов
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].auto_size = True
    
    async def _add_statistics_sheet(self, wb: Workbook, header_font: Font, header_fill: PatternFill):
        """Добавляет лист со статистикой"""
        ws = wb.create_sheet("Статистика")
        
        # Общая статистика
        row = 1
        ws.cell(row=row, column=1, value="ОБЩАЯ СТАТИСТИКА").font = Font(bold=True, size=14)
        row += 2
        
        # Получаем статистику
        total_users = await UserRepository.get_total_count()
        total_teams = await TeamRepository.get_total_teams()
        total_tournaments = await TournamentRepository.get_total_tournaments()
        active_tournaments = await TournamentRepository.get_active_tournaments_count()
        
        stats = [
            ("Всего пользователей", total_users),
            ("Всего команд", total_teams),
            ("Всего турниров", total_tournaments),
            ("Активных турниров", active_tournaments)
        ]
        
        for stat_name, stat_value in stats:
            ws.cell(row=row, column=1, value=stat_name)
            ws.cell(row=row, column=2, value=stat_value)
            row += 1
        
        # Автоширина столбцов
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 15


# Экземпляр сервиса для использования в хэндлерах
export_service = ExportService()